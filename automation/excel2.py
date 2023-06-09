# openpyxl 라이브러리 가져오기
from openpyxl import Workbook, load_workbook

# 새로운 워크북(Excel 파일) 만들기
workbook = Workbook()
worksheet = workbook.active

# 데이터 쓰기
worksheet['A1'] = '이름'
worksheet['B1'] = '나이'
worksheet['C1'] = '성별'

worksheet.append(['John', 25, '남성'])
worksheet.append(['Emily', 30, '여성'])
worksheet.append(['Michael', 35, '남성'])
worksheet.append(['Jessica', 28, '여성'])
worksheet.append(['David', 32, '남성'])
worksheet.append(['Sarah', 27, '여성'])
worksheet.append(['Daniel', 31, '남성'])
worksheet.append(['Olivia', 29, '여성'])
worksheet.append(['Matthew', 33, '남성'])
worksheet.append(['Sophia', 26, '여성'])

# Excel 파일 저장
workbook.save('example.xlsx')

# Excel 파일 열기
loaded_workbook = load_workbook('example.xlsx')
loaded_worksheet = loaded_workbook.active

# 데이터 읽기
for row in loaded_worksheet.iter_rows(min_row=1, max_row=11, values_only=True):
    print(row)

# Excel 파일 닫기
loaded_workbook.close()
