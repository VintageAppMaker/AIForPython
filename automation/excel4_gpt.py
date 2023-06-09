# excel3_gpy.py 소스를 학습시키고 다음과 같은 프롬프트를 던졌다.

# 1. 위의 소스에서 Excel에 해더를 넣을 것입니다. 
# 2. 1번째 해더는 이름, 두번재 해더는 나이, 세번째 해더는 성별입니다. 
# 3. 해더의 배경은 모두 회색이며 폰트크기는 18입니다. 
# 소스를 만들어주세요

# 1차 생성 후, 소스가 정리가 되지않았다. 그래서 다음과 같은 프롬프트를 던졌다. 

# 소스를 정리해서 py로 만들어주세요

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font

workbook = Workbook()
worksheet = workbook.active

# 해더 추가
header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
header_font = Font(size=18)

worksheet['A1'] = '이름'
worksheet['B1'] = '나이'
worksheet['C1'] = '성별'

# 스타일 적용
worksheet['A1'].fill = header_fill
worksheet['A1'].font = header_font

worksheet['B1'].fill = header_fill
worksheet['B1'].font = header_font

worksheet['C1'].fill = header_fill
worksheet['C1'].font = header_font

worksheet.append(['John', 25, '남성'])
worksheet.append(['Emily', 30, '여성'])
worksheet.append(['Michael', 35, '남성'])
worksheet.append(['Jessica', 28, '여성'])
worksheet.append(['David', 32, '남성'])
worksheet.append(['Sarah', 27, '여성'])
worksheet.append(['Daniel', 31, '남성'])
worksheet.append(['Olivia', 29, '여성'])
worksheet.append(['Matthew', 33, '남성'])
worksheet.append(['Sophia', 26, '여성'])

# 평균 계산을 위해 나이 데이터를 저장할 리스트
ages = []

# 데이터 읽기 및 평균 계산
for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, values_only=True):
    ages.append(row[1])  # 나이 데이터를 리스트에 추가

# 평균 계산
average_age = sum(ages) / len(ages)

# 평균을 최하단에 표시
worksheet.cell(row=worksheet.max_row + 1, column=2, value='평균 나이')
worksheet.cell(row=worksheet.max_row, column=2, value=average_age)
worksheet.cell(row=worksheet.max_row, column=2).font = header_font

# Excel 파일 저장
workbook.save('example.xlsx')

# Excel 파일 열기
loaded_workbook = load_workbook('example.xlsx')
loaded_worksheet = loaded_workbook.active

# 데이터 읽기
for row in loaded_worksheet.iter_rows(min_row=1, max_row=loaded_worksheet.max_row, values_only=True):
    print(row)

# Excel 파일 닫기
loaded_workbook.close()
