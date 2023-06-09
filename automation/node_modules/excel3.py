# excel2.py의 소스를 bard에게 학습시킨다. 
# 아래와 같은 프롬프트를 요청한다. 
#위의 소스에서 
#1. 워크시트에 다음을 추가한다. 

#[Bot1', 100, '여성'], ['Bot2', 100, '남성']
#2. 나이의 평균을 최하단에 표시한다. 
#소스를 고쳐주세요

# 다음과 같은 소스코드를 얻었다. 
# 논리적 오류가 있다.

from openpyxl import Workbook, load_workbook

# 새로운 워크북(Excel 파일) 만들기
workbook = Workbook()
worksheet = workbook.active

# 데이터 쓰기
worksheet['A1'] = '이름'
worksheet['B1'] = '나이'
worksheet['C1'] = '성별'

worksheet.append(['John', 25, '남성'])
worksheet.append(['Emily', 30, '여성'])
worksheet.append(['Michael', 35, '남성'])
worksheet.append(['Jessica', 28, '여성'])
worksheet.append(['David', 32, '남성'])
worksheet.append(['Sarah', 27, '여성'])
worksheet.append(['Daniel', 31, '남성'])
worksheet.append(['Olivia', 29, '여성'])
worksheet.append(['Matthew', 33, '남성'])
worksheet.append(['Sophia', 26, '여성'])

worksheet.append(['Bot1', 100, '여성'])
worksheet.append(['Bot2', 100, '남성'])

# Excel 파일 저장
workbook.save('example.xlsx')

# Excel 파일 열기
loaded_workbook = load_workbook('example.xlsx')
loaded_worksheet = loaded_workbook.active

# 데이터 읽기
for row in loaded_worksheet.iter_rows(min_row=1, max_row=12, values_only=True):
    print(row)


# 나이의 평균을 최하단에 표시한다.
average_age = sum(row[1] for row in loaded_worksheet.iter_rows(min_row=2, max_row=12, values_only=True)) / loaded_worksheet.max_row
worksheet.cell(row=13, column=1).value = '평균 나이'
worksheet.cell(row=13, column=2).value = average_age

# Excel 파일 저장
workbook.save('example.xlsx')

# Excel 파일 닫기
loaded_workbook.close()
