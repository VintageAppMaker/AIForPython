# excel2.py의 소스를 gpt에게 학습시킨다. 
# 아래와 같은 프롬프트를 요청한다. 
#위의 소스에서 
#1. 워크시트에 다음을 추가한다. 

#[Bot1', 100, '여성'], ['Bot2', 100, '남성']
#2. 나이의 평균을 최하단에 표시한다. 
#소스를 고쳐주세요

# 다음과 같은 소스코드를 얻었다. 
# 논리적 오류가 있다.

from openpyxl import Workbook, load_workbook

workbook = Workbook()
worksheet = workbook.active

worksheet['A1'] = '이름'
worksheet['B1'] = '나이'
worksheet['C1'] = '성별'

worksheet.append(['John', 25, '남성'])
worksheet.append(['Emily', 30, '여성'])
worksheet.append(['Michael', 35, '남성'])
worksheet.append(['Jessica', 28, '여성'])
worksheet.append(['David', 32, '남성'])
worksheet.append(['Sarah', 27, '여성'])
worksheet.append(['Daniel', 31, '남성'])
worksheet.append(['Olivia', 29, '여성'])
worksheet.append(['Matthew', 33, '남성'])
worksheet.append(['Sophia', 26, '여성'])

# 평균 계산을 위해 나이 데이터를 저장할 리스트
ages = []

# 데이터 읽기 및 평균 계산
for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, values_only=True):
    ages.append(row[1])  # 나이 데이터를 리스트에 추가

# 평균 계산
average_age = sum(ages) / len(ages)

# 평균을 최하단에 표시
worksheet.cell(row=worksheet.max_row + 1, column=2, value='평균 나이')
worksheet.cell(row=worksheet.max_row, column=2, value=average_age)

# Excel 파일 저장
workbook.save('example.xlsx')

# Excel 파일 열기
loaded_workbook = load_workbook('example.xlsx')
loaded_worksheet = loaded_workbook.active

# 데이터 읽기
for row in loaded_worksheet.iter_rows(min_row=1, max_row=loaded_worksheet.max_row, values_only=True):
    print(row)

# Excel 파일 닫기
loaded_workbook.close()

